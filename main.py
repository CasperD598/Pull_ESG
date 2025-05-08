import openpyxl
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
import os
import os.path

#Create lists of the .txt files
with open("jobtext.txt", "r") as jobtext_file:
    jobtexts = jobtext_file.read()
    jobtexts = jobtexts.split("\n")

with open("header.txt", "r") as header_file:
    headers = header_file.read()
    headers = headers.split("\n")

#Create list of files in folder
def list_of_files(folder_path):
    return os.listdir(folder_path)
    
#Create list of rows where headers match
def find_match(folder_path, files, headers, jobtexts):
    #Create file to paste result
    result_wb = Workbook()

    #For each file in folder
    for file in files:
        job_row = []
        match_list = []
        data_workbook = openpyxl.load_workbook(f"{folder_path}/{file}")
        data_worksheet = data_workbook.active

        #Find job rows
        for i in range(5, data_worksheet.max_row):
            if data_worksheet.cell(row=i, column=1).value != None:
                job_row.append(i)

        #Find rows that contain header
        for x in range(5, data_worksheet.max_row):
            if data_worksheet.cell(row=x, column=1).value != None:
                for header in headers:
                    if header in str(data_worksheet.cell(row=x, column=3).value).lower():
                        if job_row.index(x)+1 < len(job_row) and [x, (job_row.index(x)+1)] not in match_list:                         
                            match_list.append([x, (job_row.index(x)+1)])
                        elif job_row.index(x)+1 > len(job_row) and [x, (job_row.index(x)+1)] not in match_list:
                            match_list.append([x, len(job_row)])

        #Search for job text in matches
        text_match = []
        for match in match_list:
            for line in range(match[0], job_row[match[1]]+1):
                for text in jobtexts:
                    if text in str(data_worksheet.cell(row=line, column=3).value).lower():
                        text_match.append([line, str(data_worksheet.cell(row=line, column=3).value).lower(), str(data_worksheet.cell(row=line, column=5).value).lower()])
        
        #Paste result
        result_wb.create_sheet(file)
        result_ws = result_wb[file]
        result_ws.cell(row=1, column=1).value = "ROW"
        result_ws.cell(row=1, column=2).value = "TEXT"
        result_ws.cell(row=1, column=3).value = "AMOUNT"
        x = 2
        for match in text_match:
            result_ws.cell(row=x, column=1).value = match[0]
            result_ws.cell(row=x, column=2).value = match[1]
            result_ws.cell(row=x, column=3).value = match[2]
            x += 1

    result_wb.remove(result_wb["Sheet"])
    result_wb.save(f"{folder_path}/Resultat.xlsx")
    result_wb.close()

    if os.path.isfile(f"{folder_path}/Resultat.xlsx"):
        messagebox.showinfo(message=f"Done!\nResult was place in {folder_path}/Resultat.xlsx")  
    else:
        messagebox.showerror(message="An error has occured!")


if __name__ == "__main__":
    try:
        folder_path = filedialog.askdirectory(title="Hvor ligger filerne?")
        files = list_of_files(folder_path)
        find_match(folder_path, files, headers, jobtexts)
    except:
        pass
    
