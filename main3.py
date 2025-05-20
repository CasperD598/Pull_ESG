import openpyxl
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
import os
import os.path

with open("jobtext.txt", "r") as jobtext_file:
    jobtexts = jobtext_file.read()
    jobtexts = jobtexts.split("\n")

#Create list of files in folder
def list_of_files(folder_path):
    return os.listdir(folder_path)

def row_area(job, job_row, worksheet):
    if job_row.index(job) < len(job_row)-1:
       return job_row[job_row.index(job)+1]
    else:
       return worksheet.max_row


#Create list of rows where headers match
def find_match(folder_path, files, jobtexts):
    #Create file to paste result   
    result_wb = Workbook()
     
    #For each file in folder
    for file in files:
        match_list = []
        job_row = []

        data_workbook = openpyxl.load_workbook(f"{folder_path}/{file}")
        data_worksheet = data_workbook.active

        for i in range(5, data_worksheet.max_row):
            if data_worksheet.cell(row=i, column=1).value != None:
                job_row.append(i)

        for job in job_row:
            next_job_row = row_area(job, job_row, data_worksheet)
            for line in range(job, next_job_row):
                match_list_lenght = len(match_list)
                if data_worksheet.cell(row=line, column=3).value != None:
                    job_text = data_worksheet.cell(row=line, column=3).value.lower()
                    for text in jobtexts:
                        if text in job_text and data_worksheet.cell(row=line, column=4).value != "N/A" and data_worksheet.cell(row=line, column=4).value != "CANCELLED":
                            match_list.append([data_worksheet.cell(row=job, column=1).value, job_text, data_worksheet.cell(row=line, column=5).value])
                            break
                    if len(match_list) > match_list_lenght:
                        break               
        sorted = {
            "Oily waste":0,
            "Fresh water":0,
            "Shore power":0,
            "Sewage/Grey water":0,
            "Slop":0
        }

        #Paste result in excel
        result_wb.create_sheet(file)
        result_ws = result_wb[file]
        result_ws.cell(row=1, column=1).value = "TEXT"
        result_ws.cell(row=1, column=2).value = "AMOUNT"

        row = 2
        for line in sorted:
            result_ws.cell(row=row, column=1).value = line
            result_ws.cell(row=row, column=2).value = sorted[line]
            row += 1

        #Paste unused matches for manual check
        row += 1
        result_ws.cell(row=row, column=1).value = "REST PLEASE CHECK:"
        row += 1
        result_ws.cell(row=row, column=1).value = "JOB"
        result_ws.cell(row=row, column=2).value = "TEXT"
        result_ws.cell(row=row, column=3).value = "AMOUNT"
        row += 1

        for line in match_list:
            result_ws.cell(row=row, column=1).value = line[0]
            result_ws.cell(row=row, column=2).value = line[1]
            result_ws.cell(row=row, column=3).value = line[2]
            row += 1
    result_wb.remove(result_wb["Sheet"])
    result_wb.save(f"{folder_path}/Resultat.xlsx")
    result_wb.close()

    if os.path.isfile(f"{folder_path}/Resultat.xlsx"):
        messagebox.showinfo(message=f"Done!\nResult was placed in {folder_path}/Resultat.xlsx")  
    else:
        messagebox.showerror(message="An error has occured!")                 


if __name__ == "__main__":
    #try:
        folder_path = filedialog.askdirectory(title="Hvor ligger filerne?")
        files = list_of_files(folder_path)
        if "Resultat.xlsx" in files:
            files.pop(files.index("Resultat.xlsx"))
        find_match(folder_path, files, jobtexts)
    #except:
       #messagebox.showerror(message="An error has occured!")