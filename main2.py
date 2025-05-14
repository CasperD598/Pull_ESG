import openpyxl
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
import os
import os.path

#Create lists of the .txt files
with open("jobtext.txt", "r") as jobtext_file:
    jobtexts = jobtext_file.read()
    jobtexts = jobtexts.split("\n")

with open("header.txt", "r") as header_file:
    headers = header_file.read()
    headers = headers.split("\n")

with open("oily.txt", "r") as oily_file:
    oily = oily_file.read()
    oily = oily.split("\n")

with open("sewage.txt", "r") as sewage_file:
    sewage = sewage_file.read()
    sewage = sewage.split("\n")

with open("slop.txt", "r") as slop_file:
    slop = slop_file.read()
    slop = slop.split("\n")

with open("fresh.txt", "r") as fresh_file:
    fw = fresh_file.read()
    fw = fw.split("\n")

with open("power.txt", "r") as power_file:
    power = power_file.read()
    power = power.split("\n")

#Create list of files in folder
def list_of_files(folder_path):
    return os.listdir(folder_path)
    
#Create list of rows where headers match
def find_match(folder_path, files, headers, jobtexts):
    #Create file to paste result
    result_wb = Workbook()

    #For each file in folder
    for file in files:
        job_row = []
        match_list = []
        tank_match = []

        data_workbook = openpyxl.load_workbook(f"{folder_path}/{file}")
        data_worksheet = data_workbook.active

        #Find job rows
        for i in range(5, data_worksheet.max_row):
            if data_worksheet.cell(row=i, column=1).value != None:
                job_row.append(i)
   
        #Find rows that contain header
        for job in job_row:
            job_text = data_worksheet.cell(row=job, column=3).value
            for header in headers:
                if header in job_text and "tank" not in job_text and "tanks" not in job_text:
                    match_list.append([job, job_text])
                    break
                elif "tank" in job_text or "tanks" in job_text:
                    tank_match.append([job, job_text])
                    break
    
        results = []
        for match in match_list:
            next_job_row = job_row[job_row.index(match[0]) + 1]
            for line in range(match[0], next_job_row):
                for text in jobtexts:
                    if text in str(data_worksheet.cell(row=line, column=3).value).lower() and data_worksheet.cell(row=line, column=5).value != None and data_worksheet.cell(row=line, column=5).value != 0:
                        results.append([line, data_worksheet.cell(row=match[0], column=1).value, data_worksheet.cell(row=line, column=3).value, data_worksheet.cell(row=line, column=5).value])
                        break
        for match in tank_match:
            next_job_row = job_row[job_row.index(match[0]) + 1]
            for line in range(match[0], next_job_row):
                for text in jobtexts:
                    if text in str(data_worksheet.cell(row=line, column=3).value).lower() and data_worksheet.cell(row=line, column=5).value != None and data_worksheet.cell(row=line, column=5).value != 0:
                        results.append([line, data_worksheet.cell(row=match[0], column=1).value, data_worksheet.cell(row=line, column=3).value, data_worksheet.cell(row=line, column=5).value])
                        break
        sorted = {
            "Oily waste":0,
            "Fresh water":0,
            "Shore power":0,
            "Sewage/Grey water":0,
            "Slop":0
        }
        used_result = []
        oily_row = []
        for result in results:
            for line in fw:
                if line in str(result[2]).lower():
                    sorted["Fresh water"] = sorted["Fresh water"] + result[3]
                    used_result.append(result)
                    break
            for line in power:
                if line in str(result[2]).lower():
                    sorted["Shore power"] = sorted["Shore power"] + result[3]
                    used_result.append(result)
                    break
            for line in sewage:
                if line in str(result[2]).lower():
                    sorted["Sewage/Grey water"] = sorted["Sewage/Grey water"] + result[3]
                    used_result.append(result)
                    break
            for line in slop:
                if line in str(result[2]).lower():
                    sorted["Slop"] = sorted["Slop"] + result[3]
                    used_result.append(result)
                    break
            for line in oily:
                if line in str(result[2]).lower():      
                    sorted["Oily waste"] = sorted["Oily waste"] + result[3]
                    used_result.append(result)
                    oily_row.append(result[0])
                    break
                elif "m3 of oily pumpable waste" in str(result[2]).lower():
                    sorted["Oily waste"] = sorted["Oily waste"] + 2
                    used_result.append(result)
                    oily_row.append(result[0])
                    break
                elif "possible additional disposal, each m3" in str(result[2]).lower():
                    sorted["Oily waste"] = sorted["Oily waste"] + result[3]
                    used_result.append(result)
                    oily_row.append(result[0])
                    break
        print(sorted)
        for result in used_result:
            if result in results:
                results.pop(results.index(result))


        result_wb.create_sheet(file)
        result_ws = result_wb[file]
        result_ws.cell(row=1, column=1).value = "TEXT"
        result_ws.cell(row=1, column=2).value = "AMOUNT"

        row = 2
        for line in sorted:
            result_ws.cell(row=row, column=1).value = line
            result_ws.cell(row=row, column=2).value = sorted[line]
            row += 1

        result_ws.cell(row=row, column=1).value = "REST PLEASE CHECK:"
        row += 1
        result_ws.cell(row=row, column=1).value = "ROW"
        result_ws.cell(row=row, column=2).value = "TEXT"
        result_ws.cell(row=row, column=3).value = "AMOUNT"
        row += 1
        for line in results:
            result_ws.cell(row=row, column=1).value = line[0]
            result_ws.cell(row=row, column=2).value = line[2]
            result_ws.cell(row=row, column=3).value = line[3]
            row += 1
    
    result_wb.remove(result_wb["Sheet"])
    result_wb.save(f"{folder_path}/Resultat.xlsx")
    result_wb.close()

    print(oily_row)

    if os.path.isfile(f"{folder_path}/Resultat.xlsx"):
        messagebox.showinfo(message=f"Done!\nResult was placed in {folder_path}/Resultat.xlsx")  
    else:
        messagebox.showerror(message="An error has occured!")


if __name__ == "__main__":
    #try:
        folder_path = filedialog.askdirectory(title="Hvor ligger filerne?")
        files = list_of_files(folder_path)
        find_match(folder_path, files, headers, jobtexts)
    #except:
       #messagebox.showerror(message="An error has occured!")
    
